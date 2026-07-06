"""
export_excel.py -- generate the daily Excel snapshot.
"""
import pandas as pd
from datetime import date
from pathlib import Path

BASE      = Path(__file__).parent
XLSX_DIR  = BASE / "excel"

COLUMNS = [
    ("symbol",          "Symbol"),
    ("name",            "Name"),
    ("price",           "Price"),
    ("change",          "Change"),
    ("change_pct",      "Change %"),
    ("volume",          "Volume"),
    ("market_cap",      "Market Cap"),
    ("pe_ratio",        "P/E Ratio (TTM)"),
    ("wk52_change_pct", "52 Wk Change %"),
]

COL_FORMATS = {
    "Price":           '$#,##0.00',
    "Change":          '+$#,##0.00;-$#,##0.00',
    "Change %":        '+0.00"%";-0.00"%"',
    "Volume":          '#,##0',
    "Market Cap":      '$#,##0,,,"B"',
    "P/E Ratio (TTM)": '0.00',
    "52 Wk Change %":  '+0.00"%";-0.00"%"',
}

# columns that get green/red fill based on positive/negative value
COLOR_COLS = {"Change", "Change %", "52 Wk Change %"}

GREEN_FONT = "006100"
GREEN_FILL = "C6EFCE"
RED_FONT   = "9C0006"
RED_FILL   = "FFC7CE"


def _format_sheet(ws, n_rows):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F3864")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col_idx, (_, label) in enumerate(COLUMNS, 1):
        fmt   = COL_FORMATS.get(label)
        color = label in COLOR_COLS

        for row in range(2, n_rows + 2):
            cell = ws.cell(row=row, column=col_idx)
            if fmt:
                cell.number_format = fmt
            if color and cell.value is not None:
                try:
                    v = float(cell.value)
                    if v > 0:
                        cell.font = Font(color=GREEN_FONT, bold=True)
                        cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
                    elif v < 0:
                        cell.font = Font(color=RED_FONT, bold=True)
                        cell.fill = PatternFill("solid", fgColor=RED_FILL)
                except (TypeError, ValueError):
                    pass

        ws.column_dimensions[get_column_letter(col_idx)].auto_size = True


def export_excel(conn):
    today = date.today().isoformat()
    XLSX_DIR.mkdir(exist_ok=True)
    XLSX_PATH = XLSX_DIR / f"stocks_{today}.xlsx"
    df = pd.read_sql_query("""
        SELECT dp.symbol, s.name,
               dp.price, dp.change, dp.change_pct,
               dp.volume, dp.market_cap, dp.pe_ratio, dp.wk52_change_pct
        FROM   daily_prices dp
        JOIN   stocks s ON s.symbol = dp.symbol
        WHERE  dp.date = ? AND dp.source = 'most_active'
        ORDER  BY dp.volume DESC NULLS LAST
    """, conn, params=(today,))

    if df.empty:
        print("  No data yet -- skipping Excel export.")
        return

    df.columns = [label for _, label in COLUMNS]

    with pd.ExcelWriter(str(XLSX_PATH), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Most Active (Today)", index=False)
        _format_sheet(writer.sheets["Most Active (Today)"], len(df))

    print(f"  Excel saved -> {XLSX_PATH}")
